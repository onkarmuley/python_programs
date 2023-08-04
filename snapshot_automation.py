import openpyxl

   
def main():
    
    s_system = input("Enter source system (SAP/AX2019/AX2012):")
    path = input("Enter file path:") #"C:\\Users\\onkar_muley\\Downloads\\StagingTablesMapping.xlsx"
    snapshot_name = input("Enter snapshot name:")
    landing_table_name = input("Enter landing table name:")
    min_row = int(input("Enter row number where business column starts:"))
    
    
    #opening excel
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.get_sheet_by_name(landing_table_name.upper())
    max_row = sheet_obj.max_row
    #print("max rows", max_row)
    
    
    if s_system == 'SAP':
        sap_snapshot(sheet_obj,snapshot_name,landing_table_name,min_row,max_row)
    else:
        print("Invalid source system")
    
def sap_snapshot(sheet_obj,snapshot_name,landing_table_name,min_row,max_row):
    
    
    sap_col_list, sap_alias_list = [], []
    for i in range(min_row, max_row):
        #print(sheet_obj.cell(i,1).value.lower())
        sap_col_list.append(sheet_obj.cell(i,1).value.lower())
        sap_alias_list.append(sheet_obj.cell(i,3).value.lower())

    #print(sheet_obj)
    #print("Row starts from", min_row, max_row)
    #print(sap_col_list)
    #print(sap_alias_list)
    #print(len(sap_col_list), " ", len(sap_alias_list))

    check_cols = ""
    select_cols = ""
    for i in range(len(sap_alias_list)):
        check_cols = check_cols + "\t\t\t\t\t'" + sap_alias_list[i] + "'" if i > 0 else check_cols + "'" + sap_alias_list[i] + "'"
        select_cols = select_cols + sap_col_list[i] + " AS " + sap_alias_list[i]
        if i != len(sap_alias_list)-1:
            check_cols = check_cols + ",\n"
            select_cols = select_cols + ",\n"
    #print(len(check_cols))
    #print(select_cols)
    
    snapshot = "{% snapshot " + snapshot_name + " %}\n\n"
    
    #getting key columns 
    business_key = sheet_obj.cell(2,2).value.lower().replace(" ","").split(",")
    business_key1 = ""
    for i in range(len(business_key)):
        business_key1 = business_key1 + "'" + business_key[i] + "'"
        if i != len(business_key)-1:
            business_key1 = business_key1 + ","
    
    #getting source systems 
    source_sys = sheet_obj.cell(3,2).value.lower().replace(" ","").split(",")
    
    #preparing snapshot 
    snapshot = snapshot + "{% set business_key=[" + business_key1 + "] %}\n\n"
    
    temp_a = ""
    temp_b = ""
    temp_pre_hook = ""
    temp_post_hook = ""
    temp_select = ""
    soft_delete = "\"{{ mco_soft_delete_sap('" + landing_table_name +"',sap_data_source_cd_list=[\n"
           
    count = 0
    
    for i in source_sys:
        snapshot = snapshot + "{% set sap_" + i + "_full_criteria = mco_sap_snapshot_criteria(var('env_source_cd_sap'),var('data_source_cd_sap_" + i + "'),var('source_system'),var('run_type_full'),var('run_type'))  %} \n"
        temp_a = temp_a + "{% set sap_" + i + "_business_key =  mco_sap_surrogate_key(var('env_source_cd_sap'),var('data_source_cd_sap_" + i + "'),business_key)  %}\n"
        temp_b = temp_b + "{% set sap_" + i + "_business_key =  [var('env_source_cd_sap'),var('data_source_cd_sap_" + i + "')] + business_key  %}\n"
        temp_pre_hook = temp_pre_hook + "\"UPDATE {{ source('sap_" + i + "','" + landing_table_name + "') }} SET STAGE_LOAD_STATE_CD = 'I' WHERE STAGE_LOAD_STATE_CD ='N' AND {{ mco_sap_snapshot_criteria(var('env_source_cd_sap'),var('data_source_cd_sap_" + i + "'),var('source_system'),var('run_type_full'),var('run_type')) }}\"\n"
        temp_post_hook = temp_post_hook + "\"UPDATE {{ source('sap_" + i + "','" + landing_table_name + "') }} SET STAGE_LOAD_STATE_CD = 'P' WHERE STAGE_LOAD_STATE_CD ='I' AND {{ mco_sap_snapshot_criteria(var('env_source_cd_sap'),var('data_source_cd_sap_" + i + "'),var('source_system'),var('run_type_full'),var('run_type')) }}\"\n"
        temp_select = temp_select + """SELECT
    {{ var('env_source_cd_sap') }} env_src_cd,
    {{ var('data_source_cd_sap_""" + i + """') }} data_src_cd,
    {{ dbt_utils.surrogate_key(sap_pr1_business_key)}} etl_hash_bus_key,
  CAST({{ env_var('DBT_CLOUD_JOB_ID','\\'localrun\\'') }} AS VARCHAR(255)) etl_job_name,
        """ + select_cols + "\nFROM {{ source('sap_" + i + "','" + landing_table_name + """') }}
WHERE stage_load_state_cd = 'I' AND {{ sap_""" + i +"""_full_criteria }}
 QUALIFY ROW_NUMBER() OVER (PARTITION BY {{ business_key|join(',')}} ORDER BY stage_load_dtm DESC) = 1 \n """
        soft_delete = soft_delete + "{'sap_data_source_cd': var('data_source_cd_sap_" + i + "'),'sap_run_type':var('run_type_full')}"
        
        
        if count < len(source_sys)-1:
            temp_pre_hook = temp_pre_hook + ","
            temp_post_hook = temp_post_hook + ","
            temp_select = temp_select + "UNION ALL \n"
            soft_delete = soft_delete + ",\n"
        count = count + 1 
    
    soft_delete = soft_delete + "],\nbusiness_key=[" + business_key1 + "],sap_system=var('source_system')) }}\","
    
    
    snapshot = snapshot + "\n" + temp_a + "\n" + temp_b
       
    snapshot = snapshot + """{{
    config(
      target_schema='src_sap',
      alias='""" + sheet_obj.cell(1,2).value + """',
      strategy='check',
      unique_key='etl_hash_bus_key',
      check_cols= [""" + check_cols + "]," 
      
    snapshot = snapshot + "\n\n" + "pre_hook=[" + temp_pre_hook + "],"
    snapshot = snapshot + "\n\n" + "post_hook=[" + soft_delete + "\n" + temp_post_hook + "]," 
    snapshot = snapshot + '''\ntags=['src_sap']
        )
    }}'''
    snapshot = snapshot + "\n\n" + temp_select
    snapshot = snapshot + "\n\n{% endsnapshot %}"
    
    text_file = open("C:\\Users\\onkar_muley\\Downloads\\" + snapshot_name + ".sql", "w")
    n = text_file.write(snapshot)
    text_file.close()
    #print(snapshot)
    print(f'Snapshot written successfully at {0}',text_file)



if __name__ == '__main__' :
    main()
